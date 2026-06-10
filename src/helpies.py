import csv
import logging
import os
from datetime import date, datetime
from functools import wraps
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Dict, List, Tuple, Type

import openpyxl
import pandas as pd
from flask import Response, render_template, request
from sqlalchemy import asc, text
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.inspection import inspect
from sqlalchemy.orm import DeclarativeMeta
from weasyprint import HTML

ALLOWED_EXTENSIONS = {".csv", ".pdf"}

logger = logging.getLogger(__name__)

root_dir = Path(__file__).resolve().parent
COLUMN_MAP = {
    1: "id",
    2: "schueler_familienname",  # B
    3: "schueler_vorname",  # C
    4: "schueler_geschlecht",  # D
    5: "schueler_geburtsdatum",  # E
    6: "schueler_geburtsort",  # F
    7: "schueler_religion",  # G
    # 8 Teilnahme Religionsunterricht # H --> 1
    9: "schueler_staatsangehörigkeit",  # I
    10: "schueler_geburtsland",  # J
    11: "schueler_familiensprache",  # K
    12: "auslaender_einreisedatum",  # L
    13: "grundschule_datum",  # M
    # 14 (Ersteinschulungsdatum nicht feststellbar) # N --> 0
    15: "erziehungsberechtigte_1_name",  # 0
    16: "erziehungsberechtigte_1_vorname",  # P
    17: "erziehungsberechtigte_1_strasse",  # Q
    18: "erziehungsberechtigte_1_hausnummer",  # R
    19: "erziehungsberechtigte_1_PLZ",  # S
    20: "erziehungsberechtigte_1_wohnort",  # T
    21: "erziehungsberechtigte_1_stadtteil",  # U
    22: "erziehungsberechtigte_1_telefon_1",  # V
    23: "erziehungsberechtigte_1_telefon_2",  # W
    24: "erziehungsberechtigte_1_email",  # X
    25: "erziehungsberechtigte_1_art",  # Y
    # 26 Z (1. Erzb. Im Schriftverkehrverteiler) # Z
    27: "erziehungsberechtigte_2_name",  # AA
    28: "erziehungsberechtigte_2_vorname",  # AB
    29: "erziehungsberechtigte_2_strasse",  # AC
    30: "erziehungsberechtigte_2_hausnummer",  # AD
    31: "erziehungsberechtigte_2_PLZ",  # AE
    32: "erziehungsberechtigte_2_wohnort",  # AF
    33: "erziehungsberechtigte_2_stadtteil",  # AG
    34: "erziehungsberechtigte_2_telefon_1",  # AH
    35: "erziehungsberechtigte_2_telefon_2",  # AI
    36: "erziehungsberechtigte_2_email",  # AJ
    37: "erziehungsberechtigte_2_art",  # AK
    # 38 AL (2. Erzb. Im Schriftverkehrverteiler)
    39: "schueler_strasse",  # AM
    40: "schueler_hausnummer",  # AN
    41: "schueler_PLZ",  # AO
    42: "schueler_stadtteil",  # AP
    43: "schueler_telefon_1",  # AQ
    44: "schueler_telefon_2",  # AR
    45: "schueler_email",  # AS
    46: "erziehungsberechtigte_2_telefon_1",  # AT
    # 47 (Von Schule) # AU
    48: "letzte_schule_schulart",  # AV
    # 49 (besuchte Kindertagesstätte)
    50: "letzte_schule_klasse",  # AX
    # 51 Teilnahme Ganztagsunterricht/-betreuung (GTU) # AZ
    # 52 Teilnahme Ganztagsunterricht/-betreuung (GTU) 2 # BA
    # 53 Teilnahme Ganztagsunterricht/-betreuung (GTU) 3 # BB
    # 54 GTU Stundensatz 1 # BC
    # 55 GTU Stundensatz 2 # BD
    # 56 GTU Stundensatz 3 # BE
    # 57 Teilnahme Essen # BF
    # 58 darf Schulgelände verlassen # BG --> 1
    # 59 Interesse am herkunftssprachlichen Ergänzungsunterricht (HSU) # BH
    # 60 Veröffentlichung von Fotos BI
    # 61 Teilnahme Schulbuchausleihe BJ
    # 62 BK
    # 63 BL
    65: "ausbildung_beruf",  # BM
    66: "betrieb_name",  # BN
    67: "betrieb_anprechpartner_name",  # BO
}


# Standardwerte (Default-Werte) für bestimmte Spaltennummern
COLUMN_DEFAULTS = {
    8: 1,  # H (Teilnahme Religionsunterricht)
    14: 0,  # N (Ersteinschulungsdatum nicht feststellbar
    58: 1,  # BG (darf Schulgelände verlassen)
}

# Welche Model-Attribute sind Datumsfelder?
COLUMN_DATES = [
    "schueler_geburtsdatum",
    "eintrittsdatum",
    "grundschule_datum",
    "auslaender_einreisedatum",
]


# ------------------------------------------------------------------------------


def _init_db(state) -> None:
    """Initialisiert die SQLite-Datenbank beim App-Start, falls noch nicht vorhanden.
    Es werden, wenn nicht vorhanden, alle Tabellen erstellt. Im Besonderen wird eine Zeile
    der ConfigSetting Tabelle erstellt, um die Standardwerte für den Admin-Zugang

    Args:
        app (Flask): Flask-Applikation (wird für `app.instance_path` und `app.app_context()` benötigt).

    Returns:
        None
    """
    global STATE
    STATE = state
    try:
        # Instance-Ordner sicherstellen (muss vor Zugriff auf Datei geschehen)
        Path(STATE.app.instance_path).mkdir(parents=True, exist_ok=True)

        # Import hier, damit Modelle registriert sind, bevor create_all() aufgerufen wird
        import src.models  # noqa: F41

        try:
            STATE.db.create_all()
            config = src.models.ConfigSetting()
            state.db.session.add(config)  # Config Model mit Default Werten erstellen
            state.db.session.commit()
            logger.info("Datenbanktabellen erstellt/überprüft.")
        except SQLAlchemyError as sqle:
            logger.exception("Fehler beim Erstellen der Datenbanktabellen: %s", sqle)
            raise

    except Exception as e:
        # Globales Fehler-Logging, App-Start nicht unbedingt abbrechen, aber Hinweis geben
        logger.exception(f"_init_db -> Fehler bei der Datenbankinitialisierung: {e}")
        # Optional: weiter hochwerfen, wenn du das Starten bei Fehlern verhindern willst:
        # raise


def _update_app() -> None:
    """Lädt dynamische Konfigurationswerte aus der Datenbank und aktualisiert die Flask-App-Konfiguration.
    Diese Funktion erwartet, dass ein gültiger Application Context aktiv ist.
    Die Attributnamen des Models werden in Großbuchstaben geändert und in app.config gespeichert:
    admin_login -> app.config[ADMIN_LOGIN] = value

    Returns:
        None
    """

    def __to_dict(obj) -> Dict:
        """Wandelt den Inhalt der Tabelle in ein Dictionary um"""
        mapper = inspect(obj).mapper
        return {c.key: getattr(obj, c.key) for c in mapper.columns}

    from src.models import ConfigSetting

    try:
        for cfg in STATE.db.session.query(ConfigSetting):
            data = __to_dict(cfg)
            data.pop("id")
            for key, value in data.items():
                STATE.app.config[key.upper()] = value

    except Exception as e:
        logger.exception(f"Konnte App-Konfiguration nicht aus DB laden: {e}")


def _export_to_xlsx(schueler_liste: List[Dict[str, Any]], prototyp: Path, sheet_name: str = "Bewerber") -> BytesIO:
    try:
        if schueler_liste is None:
            raise ValueError("Schueler_liste darf nicht None sein")
        if not prototyp.exists():
            raise FileNotFoundError(f"Prototyp-Datei nicht gefunden: {prototyp}")

        # 1. Den Prototyp (die Excel-Vorlage) laden
        # data_only=True ignoriert eventuelle Formeln in der Vorlage
        workbook = openpyxl.load_workbook(prototyp, data_only=True)

        # Das richtige Sheet auswählen (oder das erste aktive, falls Name nicht existiert)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        # 2. Startzeile für die Daten ermitteln
        # Weil Zeile 1 die Header sind, fangen wir in Zeile 2 an zu schreiben
        start_row = 2

        # 3. Schülerliste durchlaufen und Daten eintragen
        for row_idx, schueler in enumerate(schueler_liste, start=start_row):
            # Jede Spalte durchgehen, die im COLUMN_MAP oder COLUMN_DEFAULTS definiert ist
            # Wir nutzen ein Set aus allen betroffenen Spaltennummern
            all_target_columns = set(COLUMN_MAP.keys()).union(COLUMN_DEFAULTS.keys())

            for col_idx in all_target_columns:
                val = None

                # A: Wert aus dem Schüler-Dictionary holen (falls gemappt)
                model_attr = COLUMN_MAP.get(col_idx)
                if model_attr:
                    val = schueler.get(model_attr)

                    # Datumsformatierung anwenden
                    if model_attr in COLUMN_DATES and val:
                        # Falls es ein String ist, zu datetime wandeln (passe das Format ggf. an)
                        if isinstance(val, str):
                            try:
                                val = datetime.strptime(val, "%Y-%m-%d")  # Angenommenes Quellformat
                            except ValueError:
                                pass  # Wenn Konvertierung fehlschlägt, den String lassen

                        # openpyxl schreibt echte datetime-Objekte nativ als Excel-Datum
                        if isinstance(val, datetime):
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            cell.value = val
                            cell.number_format = "DD.MM.YYYY"
                            continue

                # B: Wenn kein Wert da ist, Default-Wert prüfen
                if val is None or val == "":
                    val = COLUMN_DEFAULTS.get(col_idx, "")

                # Wert in die Zelle schreiben
                sheet.cell(row=row_idx, column=col_idx, value=val)

        # 4. Datei in den BytesIO-Speicher schreiben
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        # Pointer zurücksetzen
        output.seek(0)
        return output

    except ValueError as e:
        logger.error(f"_export_to_xlsx -> {e}")
        return False
    except Exception as e:
        logger.error(f"_export_to_xlsx -> Allgemeiner Fehler: {e}")
        return False


def _export_to_pdf(schueler_liste: List[Dict[str, Any]], schulform: str, klasse: str = "Alle Klassen") -> BytesIO:
    """Erzeugt ein PDF mit allen übergebenen Schülerdatensätzen und liefert es als BytesIO zurück.

    Die Funktion rendert zunächst ein HTML-Template mit Titel/Untertitel (aktuelles Schuljahr)
    und der Schülerliste. Anschließend wird das HTML mit WeasyPrint in ein PDF konvertiert und
    als BytesIO zurück gesendet

    Args:
        alle_schueler (list | tuple): Sammlung der zu exportierenden Schülerdatensätze,
            typischerweise eine Liste von ORM-Objekten oder Dictionaries, die das Template erwartet.
        schulform (str): Schlüssel/Bezeichnung der Schulform (z. B. "berufsschule", "vollzeitschule");
            wird für Titel/Dateinamen verwendet.
        klasse (str): Klassennamen

    Returns:
        BytesIO: Inhalt der zu sendenden Datei

    Hinweise:
        - Das Template "pdf_layout.html" muss die Variablen "schueler", "titel" und "untertitel" erwarten.
        - WeasyPrint führt kein JavaScript aus; eingebettete Links bleiben klickbar.
        - Für konsistente Jahresdarstellung wird das Schuljahr als "YYYY/YYYY+1" formatiert.
    """
    titel = schulform.capitalize()

    # HTML aus Template erzeugen
    try:
        html_content = render_template(
            "pdf_layout.html",
            schueler=schueler_liste,
            titel=titel,
            untertitel=f"{date.today().year}/{date.today().year + 1}",
            klasse=klasse,
        )
        # PDF-Daten generieren
        pdf_io = BytesIO()
        HTML(string=html_content).write_pdf(target=pdf_io)
        pdf_io.seek(0)
        return pdf_io

    except Exception as e:
        logger.error(f"_export_to_pdf -> PDF‑Erstellung fehlgeschlagen: {e}")
        return False


def _export_to_csv(schueler_liste: List[Dict[str, Any]]) -> BytesIO:
    """Erstellt eine CSV-Datei aus einer Liste von Schüler-Dictionaries und gibt den Inhalt als BytesIO zurück.

    Die Funktion schreibt eine CSV-Datei mit Semikolon als Trenner und utf-8-Encoding.
    Falls `schueler_liste` leer ist, wird eine leere Datei mit nur Headern (sofern `fields` angegeben)
    oder eine leere Datei angelegt.

    Args:
        schueler_liste (list[dict]): Liste von Dictionaries (gleiche Schlüssel pro Eintrag) mit Schülerdaten.

    Returns:
        BytesIO: Inhalt der zu sendenden Datei
    """

    try:
        if schueler_liste is None:
            raise ValueError("schueler_liste darf nicht None sein")

        # DataFrame erstellen – bei leerer Liste leeres DF mit Columns (falls übergeben)
        df = pd.DataFrame.from_records(schueler_liste or [])

        # NaN -> leerer String (optional)
        df = df.fillna("")

        str_io = StringIO()

        # Pandas schreibt Text in StringIO
        df.to_csv(str_io, sep=";", index=False, encoding="utf-8", lineterminator="\n")

        csv_bytes = str_io.getvalue().encode("utf-8")
        str_io.close()
        bytes_io = BytesIO(csv_bytes)
        bytes_io.seek(0)
        return bytes_io

    except ValueError as e:
        logger.exception(f"_export_to_csv -> {e}")
        return False
    except IOError as e:
        logger.exception(f"_export_to_csv -> {e}")
        return False


def _import_klassen_from_csv(path: str) -> List[Tuple[str, str, str]]:
    """_Liest die CSV‑Datei ein und liefert eine Liste von Tupeln
    (Kurzform, Bezeichnung, Art).

    Die Datei eine Kopfzeile besitzen, das Format haben:
        Kurzform;Bezeichnung;Art
    Args:
        path (str): Pfad zur CSV-Datei mit Schulklassen

    Raises:
        FileNotFoundError: _description_
        ValueError: _description_
        FileNotFoundError: _description_
        RuntimeError: _description_

    Returns:
        List[Tuple[str, str, str]]: _description_
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"CSV-Datei nicht gefunden: {path}")

    klassen: List[Tuple[str, str, str]] = []
    try:
        with open(path, mode="r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f, delimiter=";")
            # Kopfzeile überspringen
            header = next(reader, None)
            if header is None or len(header) < 3:
                raise ValueError(f"Ungültiger Header in {path}. Erwartet mind. 3 Spalten.")

            for i, row in enumerate(reader, start=2):
                if len(row) < 3:
                    # Zeile ignorieren, wenn nicht genügend Spalten vorhanden
                    logger.warning("Zeile %d in %s hat <3 Spalten, übersprungen.", i, path)
                    continue

                kurzform, bezeichnung, art = (cell.strip() for cell in row[:3])
                if not kurzform or not bezeichnung or not art:
                    logger.warning("Zeile %d in %s unvollständig, übersprungen.", i, path)

                klassen.append((kurzform, bezeichnung, art))

    except FileNotFoundError as e:
        raise FileNotFoundError(f"_import_klassen_from_csv -> CSV‑Datei nicht gefunden: {path}") from e
    except Exception as e:
        raise RuntimeError(f"_import_klassen_from_csv -> Fehler beim Einlesen der CSV: {path}") from e

    return klassen


def _has_same_header(file_1: str, file_2: str) -> bool:
    """vergleicht den Header zweier CSV-Dateien

    Args:
        file_1 (str): Pfad zur ersten Datei
        file_2 (str): Pfad zur ersten Datei

    Returns:
        bool: True, wenn beide Header identisch sind
    """
    df_header_1 = pd.read_csv(file_1, nrows=0)
    df_header_2 = pd.read_csv(file_2, nrows=0)
    return df_header_1.columns == df_header_2.columns


def _get_error_page(error: str = "") -> tuple:
    """liefert eine Fehlerseite zurück, die als Antwort auf fehlerahfte Webanfragen dient
    Die Fehlermeldung wird angezeigt

    Returns:
    tuple:  (rudimentäre Webseite, 400)
    """
    return ("bad request!", 400)


def _get_klassen_choices(schulform) -> list:
    """liefert eine Liste mit Tuples, die als Auswahl für ein PullDown-Menü Dateinamen
    Das erste Tuple besteht aus ("", "Bitte wählen..."). Alle anderen aus (Kurzform, Bezeichnung)

    Args:
        schulformen (Dictionary): Infos zu Schulformen
            name (str): Name der Schulform
            pfad (str): Pfad zur Datei mit den Klassen fürs neue Schuljahr
            liste (list): Liste mit den Klassen fürs neue Schuljahr
            model (db.Model): Das Model der jeweiligen Schulform

    Returns:
        list: Liste mit Tuples
    """
    return [
        ("", "Bitte wählen..."),
        ("NN", "UNBEKANNT"),
    ] + [(kurz, f"{bezeichnung} ({kurz})") for kurz, bezeichnung in schulform.liste]


def _save_file(file, file_path: str) -> Tuple[str, str]:
    """Speichert ein hochgeladenes Dateiobjekt auf dem Dateisystem.

    Args:
        file: Dateiupload-Objekt (z. B. request.files["file"])
        file_path: kompletter Zielpfad inklusive Dateiname.

    Returns:
        Tuple[str, str]: (Nachricht, Kategorie) mit Kategorie in {"success", "warning", "error"}

    Raises:
        IOError/OSError: Bei Schreibfehlern auf dem Dateisystem.
    """

    # Typisierte Path-Objekte
    target_path = Path(file_path).resolve()

    # Sicherstellen, dass es den Ordner für die Datei gibt
    try:
        target_path.parent.mkdir(parents=True, exist_ok=True)

    except OSError as e:
        logger.error(f"Fehler beim Erstellen des Zielordners: {e}")
        return ("Zielordner kann nicht erstellt werden.", "error")

    # Datei speichern
    try:
        file.save(target_path)
        return (f'Datei "{file.filename}" erfolgreich hochgeladen.', "success")

    except Exception as e:
        logger.error(f"_save_file -> Fehler beim Speichern der Datei: {e}")
        return ("Die Datei konnte aufgrund eines Serverfehlers nicht gespeichert werden.", "error")


def _get_schueler_list(db, model_cls: Type[DeclarativeMeta], schulklasse: str) -> List[Dict]:
    """Erstellt eine Liste von Dictionaries für Schüler der angegebenen Klasse oder aller Klassen.

    Args:
        model_cls (Type): SQLAlchemy-Modelklasse (z. B. Berufsschueler, Vollzeitschueler).
        schulklasse (str): Klassenbezeichnung oder "all" für alle Klassen.

    Returns:
        List[Dict]: Liste mit Dictionaries je Schüler (ohne SQLAlchemy-Metadaten).
    """
    MAX_EXPORT_ROWS = 1000

    try:
        # Query: alle Spalten
        stmt = db.select(*model_cls.__table__.columns).order_by(
            asc(model_cls.klassenbezeichnung),
            asc(model_cls.schueler_familienname),
        )

        # Optionaler Klassenfilter
        if schulklasse and schulklasse.lower() != "all":
            stmt = stmt.where(model_cls.klassenbezeichnung == schulklasse)

        # Hard-Limit in SQL, um große Resultsets früh zu kappen
        stmt = stmt.limit(MAX_EXPORT_ROWS + 1)

        result = db.session.execute(stmt)
        records = [dict(r) for r in result.mappings()]  # List[dict]
        return records

    except SQLAlchemyError as e:
        logger.error(f"Datenbankfehler in _get_schueler_list: {e}")
        return []
    except Exception as e:
        logger.error(f"Fehler in _get_schueler_list: {e}")
        return []


# + ----------------------------------------------------------------------------
# + AUTHENTIFIZIERUNG
# + ----------------------------------------------------------------------------
def __check_auth(username, password) -> bool:
    """Überprüft, ob username und Passwort stimmen

    Args:
        username (str): Benutzername, der überprüft werden soll
        password (str): Passwort, das überprüft werden soll

    Returns:
        bool: True, wenn Beide mit den erlaubten übereinstimmen
    """
    return username == STATE.app.config["ADMIN_LOGIN"] and password == STATE.app.config["ADMIN_PASSWORD"]


def __authenticate():
    """HTML Anfrage zur Authentifizierung"""
    return Response(
        "Login erforderlich",
        401,
        {"WWW-Authenticate": 'Basic realm="Login erforderlich"'},
    )


def _requires_auth(f):
    """Überprüft ob eine Authentifizierung notwendig ist oder bereits durchgeführt wurde"""

    @wraps(f)
    def decorated(*args, **kwargs):

        auth = request.authorization
        if not auth or not __check_auth(auth.username, auth.password):
            return __authenticate()
        return f(*args, **kwargs)

    return decorated


# ------------


def update_db():
    """Ändert die Datenbankstruktur
    update_db() muss in _initdb() nach STATE.db.create_all() ausgeführt werden
    """
    print(1)
    STATE.db.session.execute(text("ALTER TABLE Vollzeitschueler ADD COLUMN erziehungsberechtigte_1_art VARCHAR(30)"))
    STATE.db.session.execute(
        text("UPDATE Vollzeitschueler SET erziehungsberechtigte_1_art = 'Mu' WHERE erziehungsberechtigte_1_art IS NULL")
    )
    print(2)
    STATE.db.session.execute(text("ALTER TABLE Vollzeitschueler ADD COLUMN erziehungsberechtigte_2_art VARCHAR(30)"))
    STATE.db.session.execute(
        text("UPDATE Vollzeitschueler SET erziehungsberechtigte_2_art = 'Mu' WHERE erziehungsberechtigte_2_art IS NULL")
    )
    STATE.db.session.commit()
